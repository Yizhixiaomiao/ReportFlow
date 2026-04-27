from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


@dataclass
class WorkbookPreview:
    path: Path
    sheet_names: list[str]
    active_sheet: str
    columns: list[dict[str, Any]]
    preview_rows: list[dict[str, Any]]
    row_count: int
    column_count: int


def jsonable(value: Any) -> Any:
    if value in ("", None):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    return value


def empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def infer_type(values: list[Any]) -> str:
    real_values = [value for value in values if not empty(value)]
    if not real_values:
        return "unknown"
    if all(isinstance(value, bool) for value in real_values):
        return "boolean"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in real_values):
        return "number"
    if all(isinstance(value, (datetime, date)) for value in real_values):
        return "datetime"
    return "text"


def load_rows(path: str | Path, sheet_name: str | None = None, header_row: int = 1) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    path = Path(path)
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("仅支持 .xlsx 或 .xls 文件")
    if path.suffix.lower() == ".xls":
        return load_xls_rows(path, sheet_name, header_row)
    return load_xlsx_rows(path, sheet_name, header_row)


def load_xlsx_rows(path: Path, sheet_name: str | None, header_row: int) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet_names = workbook.sheetnames
    active_sheet = sheet_name or sheet_names[0]
    sheet = workbook[active_sheet]
    raw_rows = list(sheet.iter_rows(values_only=True))
    return rows_from_matrix(raw_rows, sheet_names, header_row)


def load_xls_rows(path: Path, sheet_name: str | None, header_row: int) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    if xlrd is None:
        raise ValueError("读取 .xls 需要安装 xlrd")
    book = xlrd.open_workbook(str(path))
    sheet_names = book.sheet_names()
    active_sheet = sheet_name or sheet_names[0]
    sheet = book.sheet_by_name(active_sheet)
    matrix = [[xls_cell_value(sheet.cell(row, col), book.datemode) for col in range(sheet.ncols)] for row in range(sheet.nrows)]
    return rows_from_matrix(matrix, sheet_names, header_row)


def xls_cell_value(cell, datemode: int) -> Any:
    if xlrd is not None and cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, datemode)
    return cell.value


def rows_from_matrix(matrix: list[tuple[Any, ...] | list[Any]], sheet_names: list[str], header_row: int) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    if len(matrix) < header_row:
        return [], [], sheet_names
    header_values = matrix[header_row - 1]
    columns = []
    seen: dict[str, int] = {}
    for index, value in enumerate(header_values, start=1):
        name = str(value).strip() if value not in (None, "") else f"列{index}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        columns.append(name)

    rows: list[dict[str, Any]] = []
    for source in matrix[header_row:]:
        row = {column: jsonable(source[index]) if index < len(source) else None for index, column in enumerate(columns)}
        if any(not empty(value) for value in row.values()):
            rows.append(row)
    return columns, rows, sheet_names


def load_workbook_preview(path: str | Path, sheet_name: str | None = None, header_row: int = 1, preview_size: int = 50) -> WorkbookPreview:
    path = Path(path)
    columns, rows, sheet_names = load_rows(path, sheet_name, header_row)
    active_sheet = sheet_name or (sheet_names[0] if sheet_names else "")
    column_infos = []
    for index, column in enumerate(columns, start=1):
        values = [row.get(column) for row in rows[:50]]
        column_infos.append(
            {
                "index": index,
                "name": column,
                "type": infer_type(values),
                "sample_values": [value for value in values if not empty(value)][:3],
            }
        )
    return WorkbookPreview(
        path=path,
        sheet_names=sheet_names,
        active_sheet=active_sheet,
        columns=column_infos,
        preview_rows=rows[:preview_size],
        row_count=len(rows),
        column_count=len(columns),
    )


def coerce_number(value: Any) -> float | None:
    try:
        if empty(value):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def compare(value: Any, operator: str, expected: Any = None) -> bool:
    if operator == "equals":
        return str(value) == str(expected)
    if operator == "not_equals":
        return str(value) != str(expected)
    if operator == "contains":
        return str(expected) in str(value)
    if operator == "not_contains":
        return str(expected) not in str(value)
    if operator == "greater_than":
        left, right = coerce_number(value), coerce_number(expected)
        return left is not None and right is not None and left > right
    if operator == "less_than":
        left, right = coerce_number(value), coerce_number(expected)
        return left is not None and right is not None and left < right
    if operator == "greater_or_equal":
        left, right = coerce_number(value), coerce_number(expected)
        return left is not None and right is not None and left >= right
    if operator == "less_or_equal":
        left, right = coerce_number(value), coerce_number(expected)
        return left is not None and right is not None and left <= right
    if operator == "is_empty":
        return empty(value)
    if operator == "is_not_empty":
        return not empty(value)
    if operator == "in":
        values = expected if isinstance(expected, list) else [item.strip() for item in str(expected).split(",")]
        return str(value) in [str(item) for item in values]
    if operator == "not_in":
        values = expected if isinstance(expected, list) else [item.strip() for item in str(expected).split(",")]
        return str(value) not in [str(item) for item in values]
    return True


def apply_cell_edits(rows: list[dict[str, Any]], rules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = [dict(row) for row in rows]
    for rule in rules:
        row_index = int(rule.get("row_index", -1))
        field = rule.get("field")
        if 0 <= row_index < len(result) and field:
            result[row_index][field] = rule.get("value")
    return result


def apply_operations(columns: list[str], rows: list[dict[str, Any]], rules: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    current_columns = list(columns)
    result = [dict(row) for row in rows]
    for rule in rules:
        action = rule.get("action")
        fields = [field for field in rule.get("fields", []) if field in current_columns]
        if action == "drop_columns":
            current_columns = [column for column in current_columns if column not in fields]
            result = [{column: row.get(column) for column in current_columns} for row in result]
        elif action == "rename_column":
            old_field = rule.get("old_field")
            new_field = rule.get("new_field")
            if old_field in current_columns and new_field:
                current_columns = [new_field if column == old_field else column for column in current_columns]
                for row in result:
                    row[new_field] = row.pop(old_field, None)
        elif action in ("add_constant_column", "add_empty_column"):
            field = rule.get("field")
            if field and field not in current_columns:
                current_columns.append(field)
                for row in result:
                    row[field] = rule.get("value") if action == "add_constant_column" else None
        elif action == "select_columns" and fields:
            current_columns = fields
            result = [{column: row.get(column) for column in current_columns} for row in result]
        elif action == "fill_empty":
            field = rule.get("field")
            if field in current_columns:
                for row in result:
                    if empty(row.get(field)):
                        row[field] = rule.get("value")
        elif action == "deduplicate":
            subset = fields or current_columns
            seen = set()
            deduped = []
            for row in result:
                key = tuple(row.get(field) for field in subset)
                if key not in seen:
                    seen.add(key)
                    deduped.append(row)
            result = deduped
    return current_columns, result


def apply_filters(rows: list[dict[str, Any]], rules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = [dict(row) for row in rows]
    for rule in rules:
        field = rule.get("field")
        result = [row for row in result if compare(row.get(field), rule.get("operator"), rule.get("value"))]
    return result


def apply_formula_fields(columns: list[str], rows: list[dict[str, Any]], rules: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    current_columns = list(columns)
    result = [dict(row) for row in rows]
    for rule in rules:
        name = rule.get("field_name")
        expression = rule.get("expression")
        if not name or not expression:
            continue
        if name not in current_columns:
            current_columns.append(name)
        for row in result:
            row[name] = eval_expression(expression, row)
    return current_columns, result


def eval_expression(expression: str, row: dict[str, Any]) -> Any:
    local_values: dict[str, Any] = {}
    rendered = expression
    for index, (field, value) in enumerate(row.items()):
        token = f"v{index}"
        local_values[token] = coerce_number(value) if coerce_number(value) is not None else value
        rendered = rendered.replace("{" + field + "}", token)
        rendered = rendered.replace(str(field), token)
    try:
        return eval(rendered, {"__builtins__": {}}, local_values)
    except Exception:
        return None


def apply_sort(rows: list[dict[str, Any]], rules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = [dict(row) for row in rows]
    for rule in reversed(rules):
        field = rule.get("field")
        result.sort(key=lambda row: (empty(row.get(field)), row.get(field)), reverse=rule.get("order", "asc") == "desc")
    return result


def apply_groups(rows: list[dict[str, Any]], rules: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    output_rows: list[dict[str, Any]] = []
    output_columns = ["汇总名称"]
    for rule in rules:
        group_fields = rule.get("group_fields", [])
        metrics = rule.get("metrics", [])
        for field in group_fields:
            if field not in output_columns:
                output_columns.append(field)
        for metric in metrics:
            alias = metric.get("alias") or f"{metric.get('field')}_{metric.get('agg')}"
            if alias not in output_columns:
                output_columns.append(alias)

        groups: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
        for row in rows:
            key = tuple(row.get(field) for field in group_fields)
            groups.setdefault(key, []).append(row)
        for key, grouped_rows in groups.items():
            output = {"汇总名称": rule.get("name") or "汇总"}
            for field, value in zip(group_fields, key):
                output[field] = value
            for metric in metrics:
                field = metric.get("field")
                agg = metric.get("agg")
                alias = metric.get("alias") or f"{field}_{agg}"
                values = [row.get(field) for row in grouped_rows if not empty(row.get(field))]
                numbers = [coerce_number(value) for value in values if coerce_number(value) is not None]
                if agg == "count":
                    output[alias] = len(values)
                elif agg == "distinct_count":
                    output[alias] = len({str(value) for value in values})
                elif agg == "sum":
                    output[alias] = sum(numbers)
                elif agg == "avg":
                    output[alias] = sum(numbers) / len(numbers) if numbers else None
                elif agg == "max":
                    output[alias] = max(numbers) if numbers else None
                elif agg == "min":
                    output[alias] = min(numbers) if numbers else None
            output_rows.append(output)
    if not output_rows:
        return ["提示"], [{"提示": "未配置汇总规则"}]
    return output_columns, output_rows


def write_sheet(workbook: Workbook, title: str, columns: list[str], rows: list[dict[str, Any]]) -> Any:
    sheet = workbook.create_sheet(title)
    for col_idx, column in enumerate(columns, start=1):
        sheet.cell(1, col_idx, column)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, column in enumerate(columns, start=1):
            sheet.cell(row_idx, col_idx, row.get(column))
    return sheet


def write_excel_formulas(sheet, columns: list[str], row_count: int, rules: list[dict[str, Any]]) -> None:
    start_col = len(columns) + 1
    for offset, rule in enumerate(rules):
        name = rule.get("field_name") or f"公式{offset + 1}"
        formula = rule.get("excel_formula") or ""
        if not formula:
            continue
        col_idx = columns.index(name) + 1 if name in columns else start_col + offset
        sheet.cell(1, col_idx, name)
        for row_idx in range(2, row_count + 2):
            rendered = formula.replace("{row}", str(row_idx))
            for source_index, column in enumerate(columns, start=1):
                rendered = rendered.replace("{" + str(column) + "}", f"{get_column_letter(source_index)}{row_idx}")
            sheet.cell(row_idx, col_idx, rendered if rendered.startswith("=") else "=" + rendered)


def apply_visual_rules(sheet, columns: list[str], rules: list[dict[str, Any]]) -> None:
    for rule in rules:
        action = rule.get("action")
        if action == "set_column_widths":
            for field, width in (rule.get("widths") or {}).items():
                if field in columns:
                    sheet.column_dimensions[get_column_letter(columns.index(field) + 1)].width = float(width)
        elif action == "set_row_heights":
            for row_index, height in (rule.get("heights") or {}).items():
                sheet.row_dimensions[int(row_index)].height = float(height)
        elif action == "set_cell_styles":
            for change in rule.get("changes") or []:
                field = change.get("field")
                if field not in columns:
                    continue
                row_index = int(change.get("row") or 1)
                col_index = columns.index(field) + 1
                style = change.get("style") or {}
                cell = sheet.cell(row_index, col_index)
                if style.get("number_format"):
                    cell.number_format = style["number_format"]
                if style.get("fill_color"):
                    cell.fill = PatternFill("solid", fgColor=style["fill_color"])
                if style.get("font_bold") or style.get("font_color"):
                    cell.font = Font(bold=bool(style.get("font_bold")), color=style.get("font_color") or None)


def apply_chart_rules(sheet, columns: list[str], row_count: int, rules: list[dict[str, Any]]) -> None:
    for index, rule in enumerate(rules, start=1):
        if row_count < 1 or not columns:
            continue
        chart_type = str(rule.get("chart_kind") or rule.get("chart_type") or "").lower()
        if "pie" in chart_type:
            chart = PieChart()
        elif "line" in chart_type:
            chart = LineChart()
        else:
            chart = BarChart()
        chart.title = rule.get("name") or f"图表{index}"
        data = Reference(sheet, min_col=2 if len(columns) > 1 else 1, max_col=min(len(columns), 2), min_row=1, max_row=row_count + 1)
        chart.add_data(data, titles_from_data=True)
        if len(columns) > 1:
            cats = Reference(sheet, min_col=1, min_row=2, max_row=row_count + 1)
            chart.set_categories(cats)
        sheet.add_chart(chart, f"{get_column_letter(len(columns) + 2)}{2 + (index - 1) * 15}")


def execute_scheme(source_path: str | Path, scheme: dict[str, Any], output_path: str | Path) -> dict[str, Any]:
    config = scheme.get("config_json", scheme)
    columns, rows, _sheet_names = load_rows(source_path, config.get("input_sheet"), int(config.get("header_row") or 1))
    rows = apply_cell_edits(rows, config.get("cell_edit_rules") or [])
    columns, rows = apply_operations(columns, rows, config.get("operation_rules") or [])
    rows = apply_filters(rows, config.get("filter_rules") or [])
    columns, rows = apply_formula_fields(columns, rows, config.get("calculated_fields") or [])
    rows = apply_sort(rows, config.get("sort_rules") or [])
    summary_columns, summary_rows = apply_groups(rows, config.get("group_rules") or [])

    workbook = Workbook()
    workbook.remove(workbook.active)
    for rule in config.get("workbook_rules") or []:
        if rule.get("action") == "add_sheet" and rule.get("sheet") and rule.get("sheet") not in workbook.sheetnames:
            workbook.create_sheet(rule["sheet"])
    detail_sheet = write_sheet(workbook, "明细数据", columns, rows)
    write_excel_formulas(detail_sheet, columns, len(rows), config.get("excel_formula_rules") or [])
    apply_visual_rules(detail_sheet, columns, config.get("visual_rules") or [])
    apply_chart_rules(detail_sheet, columns, len(rows), config.get("chart_rules") or [])
    write_sheet(workbook, "汇总数据", summary_columns, summary_rows)
    write_sheet(workbook, "异常数据", ["行号", "字段", "异常名称", "等级", "说明"], [])
    workbook.save(output_path)

    return {
        "total_rows": len(rows),
        "detail_rows": len(rows),
        "output_path": str(output_path),
        "finished_at": datetime.now().isoformat(timespec="seconds"),
    }
